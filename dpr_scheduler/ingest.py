from __future__ import annotations

import csv
from typing import List
from datetime import datetime, date, timedelta
from pathlib import Path
from openpyxl import load_workbook

from .models import Task, DPRRecord


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _read_rows_from_csv(path: str) -> List[dict]:
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return [dict({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}) for row in reader]


def _read_rows_from_xlsx(path: str) -> List[dict]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = next(rows_iter)
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[dict] = []
    for r in rows_iter:
        row_dict = {}
        for key, val in zip(headers, r):
            if key == "":
                continue
            row_dict[key] = val
        rows.append(row_dict)
    return rows


def _read_rows(path: str) -> List[dict]:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return _read_rows_from_csv(path)
    if suffix == ".xlsx":
        return _read_rows_from_xlsx(path)
    raise ValueError(f"Unsupported file extension for {path}. Supported: {SUPPORTED_EXTENSIONS}")


def _parse_dependencies(value) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return []
        parts = [p.strip() for p in s.replace(";", ",").split(",")]
        return [p for p in parts if p]
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return []


def _parse_date(value) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.fromisoformat(value).date()
    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(value))).date()
    raise ValueError(f"Unsupported date type: {type(value)}")


def read_wbs(path: str) -> List[Task]:
    rows = _read_rows(path)
    required = {"task_id", "name", "duration_days"}
    missing = required - set(rows[0].keys()) if rows else required
    if missing:
        raise ValueError(f"WBS file missing required columns: {sorted(missing)}")

    tasks: List[Task] = []
    for row in rows:
        dependencies = _parse_dependencies(row.get("dependencies"))
        duration_raw = row.get("duration_days")
        duration_days = int(float(duration_raw)) if duration_raw not in (None, "") else 0
        resource_val = row.get("resource")
        resource = str(resource_val).strip() if resource_val not in (None, "") else None
        percent_val = row.get("percent_complete", 0)
        try:
            percent = float(percent_val)
        except Exception:
            percent = 0.0
        tasks.append(
            Task(
                task_id=str(row.get("task_id", "")).strip(),
                name=str(row.get("name", "")).strip(),
                duration_days=duration_days,
                dependencies=dependencies,
                resource=resource,
                percent_complete=percent,
            )
        )
    return tasks


def read_dpr(path: str) -> List[DPRRecord]:
    rows = _read_rows(path)
    required = {"date", "task_id", "percent_complete"}
    missing = required - set(rows[0].keys()) if rows else required
    if missing:
        raise ValueError(f"DPR file missing required columns: {sorted(missing)}")

    records: List[DPRRecord] = []
    for row in rows:
        raw_date = row.get("date")
        if isinstance(raw_date, str):
            parsed_date = datetime.fromisoformat(raw_date).date()
        elif isinstance(raw_date, date):
            parsed_date = raw_date
        else:
            try:
                base = datetime(1899, 12, 30)
                parsed_date = (base + timedelta(days=int(raw_date))).date()
            except Exception:
                raise ValueError(f"Unsupported date value in DPR: {raw_date!r}")

        percent = float(row.get("percent_complete", 0))
        records.append(
            DPRRecord(
                date=parsed_date,
                task_id=str(row.get("task_id", "")).strip(),
                percent_complete=percent,
            )
        )
    return records